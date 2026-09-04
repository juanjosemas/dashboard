#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ECO STRUCT - Regenera el Excel con TODOS los datos del CSV
"""
import csv
import io
import os
import re

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_FILE = r'C:\Users\jjmax\Downloads\1\dashboard\ECO_STRUCT_-_Workspace_Gastos.xlsx'

# ============================================================
# CERTIFICACIONES - Leido desde .docx (se actualiza automaticamente)
# ============================================================
try:
    from docx import Document as _DocCert
except ImportError:
    os.system("pip install python-docx")
    from docx import Document as _DocCert

_cert_docx_path = r'C:\Users\jjmax\Downloads\1\dashboard\CERTIFICACIONES.docx'
_cert_doc = _DocCert(_cert_docx_path)

def _parse_euro_amount(s):
    """Parse European amount: handles '8875,66' and '116432.26' and '1.699,09'"""
    s = s.strip().replace('€', '').strip()
    if re.search(r',\d{1,2}$', s):
        s = s.replace('.', '').replace(',', '.')
    elif re.search(r'\.\d{1,2}$', s):
        s = s.replace(',', '')
    else:
        s = s.replace('.', '').replace(',', '')
    try:
        return float(s)
    except:
        return 0.0

cert_data = []
for _cp in _cert_doc.paragraphs:
    _ct = _cp.text.strip()
    if not _ct:
        continue
    # Find euro amounts
    _cmatches = list(re.finditer(r'([\d.,]+)\s*€', _ct))
    if not _cmatches:
        continue
    _clast = _cmatches[-1]
    _cimporte = _parse_euro_amount(_clast.group(1))
    _cnombre = _ct[:_clast.start()].strip()
    _cnombre = re.sub(r'[\s\-]+$', '', _cnombre).strip()
    _cnombre = re.sub(r'\([\d.,\s\+\-]+\)\s*=\s*$', '', _cnombre).strip()
    _cnombre = re.sub(r'[\s\-]+$', '', _cnombre).strip()
    if _cimporte > 0 and _cnombre:
        cert_data.append((_cnombre, _cimporte, ""))

total_cert = sum(c[1] for c in cert_data)
print(f'Certificaciones: {len(cert_data)} proyectos, total {total_cert:,.2f} EUR')
for i, (name, amt, obs) in enumerate(cert_data, 1):
    print(f'  {i}. {name}: {amt:,.2f} EUR')

# ============================================================
# MAPEO CSV -> CERTIFICACIONES (mismo algoritmo que el dashboard)
# ============================================================
def map_csv_to_cert(csv_name):
    """Map CSV project name to certification name using keyword matching."""
    cu = csv_name.upper()
    for cname, _, _ in cert_data:
        nu = cname.upper()
        # Direct substring match
        if nu in cu or cu in nu:
            return cname
        kw = False
        if 'FORMENTERA' in cu and 'FORMENTERA' in nu:
            if ('12' in cu and '12' in nu) or ('14' in cu and '14' in nu):
                if ('JOAQUIN' in cu and 'JOAQUIN' in nu) or ('JUANMA' in cu and 'JUANMA' in nu):
                    kw = True
        if 'EDIFICIO ALICANTE' in cu and 'ALICANTE' in nu and 'PARTE DIFICIL' in nu: kw = True
        if 'CUARTEL DE ARTILLERIA' in cu and 'MURCIA' in nu and 'PARTE DIFICIL' in nu: kw = True
        if 'EDIFICIO' in cu and 'EDIFICIO' in nu: kw = True
        if 'CUARTEL' in cu and 'CUARTEL' in nu: kw = True
        if 'SANTA ROSA' in cu and 'SANTA ROSA' in nu: kw = True
        if 'CARTAGENA' in cu and 'CARTAGENA' in nu: kw = True
        if 'CASTILLO' in cu and 'CASTILLO' in nu: kw = True
        if 'GARAJE' in cu and 'GARAJE' in nu: kw = True
        if 'PEREAMAR' in cu and 'PEREAMAR' in nu: kw = True
        if 'CBS' in cu and 'CBS' in nu: kw = True
        if 'CEMENTERIO' in cu and 'CEMENTERIO' in nu: kw = True
        if 'BARINAS' in cu and 'BARINAS' in nu: kw = True
        if 'ALBERCA' in cu and 'ALBERCA' in nu: kw = True
        if 'PLAZA CIRCULAR' in cu and 'PLAZA CIRCULAR' in nu: kw = True
        if 'HELENA' in cu and 'HELENA' in nu: kw = True
        if 'A-13' in cu and 'A-13' in nu: kw = True
        if 'CARLA' in cu and 'CARLA' in nu: kw = True
        if 'PADRE TRINI' in cu and 'PADRE TRINI' in nu: kw = True
        if 'PISO IBI' in cu and 'PISO IBI' in nu: kw = True
        if 'LUC2' in cu and 'LUC2' in nu: kw = True
        if 'MEJORA DEL VALLADO' in cu and 'CASTILLO' in nu: kw = True
        if ('ANGEL' in cu or 'NGEL' in cu) and 'HELENA' in nu and 'CARMEN' in cu: kw = True
        if kw:
            return cname
    return csv_name  # Sin mapeo: usar nombre original

def map_mo_name(mo_name):
    """Map MO name to certification name."""
    mu = mo_name.upper()
    for cname, _, _ in cert_data:
        nu = cname.upper()
        if nu in mu or mu in nu:
            return cname
        if 'FORMENTERA' in mu and 'FORMENTERA' in nu:
            if ('12' in mu and '12' in nu) or ('14' in mu and '14' in nu):
                if ('JOAQUIN' in mu and 'JOAQUIN' in nu) or ('JUANMA' in mu and 'JUANMA' in nu):
                    return cname
        if 'PARTE DIFICIL' in mu and 'PARTE DIFICIL' in nu:
            if 'ALICANTE' in mu and 'ALICANTE' in nu: return cname
            if 'MURCIA' in mu and 'MURCIA' in nu: return cname
            if 'ALICANTE' not in nu and 'MURCIA' not in nu: return cname
        if 'CEMENTERIO' in mu and 'CEMENTERIO' in nu: return cname
        if 'PEREAMAR' in mu and 'PEREAMAR' in nu: return cname
        if 'CARTAGENA' in mu and 'CARTAGENA' in nu: return cname
        if ('ANGEL' in mu or 'NGEL' in mu) and 'HELENA' in nu and 'CARMEN' in mu: return cname
    return mo_name

# ============================================================
# LEER CSV COMPLETO
# ============================================================
csv_path = r'C:\Users\jjmax\Downloads\1\dashboard\ECO_STRUCT_-_Workspace_Gastos.csv'
with open(csv_path, 'r', encoding='latin-1') as f:
    content = f.read()

reader = csv.DictReader(io.StringIO(content), delimiter=';')

gg_entries = []
vh_entries = []
gastos_directos = {}  # by project
facturas_por_obra = []  # all individual invoices (excluding GG/VEH)

for row in reader:
    proyecto = row.get('Proyecto', '').strip()
    importe_str = row.get('Importe', '0').strip()
    cod = row.get('C\u00f3digo', row.get('Codigo', '')).strip()
    fecha = row.get('Fecha de imputaci\u00f3n', '').strip()
    titulo = row.get('T\u00edtulo', '').strip()
    proveedor = row.get('Proveedor', '').strip()
    estado = row.get('Estado', '').strip()

    if importe_str:
        importe_str = importe_str.replace('.', '').replace(',', '.')
    try:
        importe = float(importe_str) if importe_str else 0
    except:
        importe = 0

    entry = {'cod': cod, 'fecha': fecha, 'titulo': titulo, 'proveedor': proveedor, 'importe': importe, 'proyecto': proyecto, 'estado': estado}

    if proyecto == 'GASTOS GENERALES':
        gg_entries.append(entry)
    elif proyecto == 'VEHICULOS':
        vh_entries.append(entry)
    elif proyecto and proyecto not in ('', 'GASTOS GENERALES', 'VEHICULOS'):
        # Map to certification names
        mapped = proyecto
        gastos_directos[mapped] = gastos_directos.get(mapped, 0) + importe
        facturas_por_obra.append(entry)

gg_total = sum(e['importe'] for e in gg_entries)
vh_total = sum(e['importe'] for e in vh_entries)

print(f'Gastos Generales: {len(gg_entries)} entradas = {gg_total:,.2f} EUR')
print(f'Vehiculos: {len(vh_entries)} entradas = {vh_total:,.2f} EUR')
print(f'Total Gastos Comunes: {gg_total + vh_total:,.2f} EUR')

# ============================================================
# APLICAR MAPEO a gastos directos del CSV
# ============================================================
gastos_directos_mapped = {}
for proj, total in gastos_directos.items():
    cert_name = map_csv_to_cert(proj)
    gastos_directos_mapped[cert_name] = gastos_directos_mapped.get(cert_name, 0) + total
gastos_directos = gastos_directos_mapped

# Also map facturas_por_obra project names
for entry in facturas_por_obra:
    entry['proyecto'] = map_csv_to_cert(entry['proyecto'])

# ============================================================
# MANO DE OBRA - Leido desde .docx (se actualiza automaticamente)
# ============================================================
try:
    from docx import Document as _DocMO
except ImportError:
    os.system("pip install python-docx")
    from docx import Document as _DocMO

_mo_docx_path = r'C:\Users\jjmax\Downloads\1\dashboard\GASTOS MANO DE OBRA.docx'
_mo_doc = _DocMO(_mo_docx_path)

import re as _re_mo

mo_data = []
_current_year = 2026  # Default year
for _mp in _mo_doc.paragraphs:
    _mt = _mp.text.strip()
    if not _mt:
        continue
    # Normalize: strip non-ASCII (euro signs, em-dashes, etc.) like dashboard does
    _mt = _re_mo.sub(r'[^\x00-\x7f]+', ' ', _mt)
    _mt = _re_mo.sub(r'\s+', ' ', _mt).strip()
    if not _mt:
        continue
    # Detect year headers
    _year_match = _re_mo.search(r'(20\d{2})', _mt)
    if 'ANO' in _mt.upper() or _mt.upper().startswith('GASTOS'):
        if _year_match:
            _current_year = int(_year_match.group(1))
        continue
    # Format 1: NAME --- HOURS HORAS A 20E - TOTAL XXXX
    _line_match = _re_mo.search(r'(\d+)\s*HORAS', _mt, _re_mo.IGNORECASE)
    if _line_match:
        _mhours = int(_line_match.group(1))
        _name_part = _mt[:_line_match.start()].strip().rstrip('-').rstrip().strip()
        _name_part = _re_mo.sub(r'[\s\-]+$', '', _name_part).strip()
        if _name_part:
            mo_data.append((_name_part, _mhours, _current_year, 0))  # cost=0 means use hours*20
        continue
    # Format 2: NAME --- TOTAL XXXX (no hours, just cost)
    _fmt2 = _re_mo.search(r'(.+?)[\s\-]+TOTAL\s*([\d.,]+)', _mt, _re_mo.IGNORECASE)
    if _fmt2:
        _name2 = _fmt2.group(1).strip().rstrip('-').strip()
        _name2 = _re_mo.sub(r'[\s\-]+$', '', _name2).strip()
        _coste2 = _fmt2.group(2).replace('.', '').replace(',', '.')
        try:
            _coste2 = float(_coste2)
        except:
            _coste2 = 0.0
        if _name2 and _coste2 > 0:
            mo_data.append((_name2, 0, _current_year, _coste2))  # 0 hours, cost explicit

if not mo_data:
    print("WARNING: No mano de obra data found in .docx")
    mo_data = []

# ============================================================
# CREAR EXCEL
# ============================================================
wb = openpyxl.Workbook()

header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=14, color='1A1A2E')
total_fill = PatternFill(start_color='E8ECF4', end_color='E8ECF4', fill_type='solid')
total_font = Font(name='Calibri', bold=True, size=11)
money_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_rows(ws, start, end, max_col, money_cols=None):
    money_cols = money_cols or []
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c in money_cols:
                cell.number_format = money_format

def style_total(ws, row, max_col, money_cols=None):
    money_cols = money_cols or []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if c in money_cols:
            cell.number_format = money_format

# --- PESTANA: Certificaciones ---
ws = wb.active
ws.title = "Certificaciones"
ws.sheet_properties.tabColor = "3498DB"
ws['A1'] = "CERTIFICACIONES A FECHA DE 31/08/2026"
ws['A1'].font = title_font
ws.merge_cells('A1:E1')
for j, h in enumerate(["N", "Proyecto", "Certificacion (EUR)", "% del Total", "Observaciones"], 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 5)
for i, (nombre, cert, obs) in enumerate(cert_data, start=4):
    ws.cell(row=i, column=1, value=i-3)
    ws.cell(row=i, column=2, value=nombre)
    ws.cell(row=i, column=3, value=cert)
    ws.cell(row=i, column=4, value=f'=C{i}/C{4+len(cert_data)}')
    ws.cell(row=i, column=5, value=obs)
tr = 4 + len(cert_data)
ws.cell(row=tr, column=2, value="TOTAL CERTIFICACIONES")
ws.cell(row=tr, column=3).value = f'=SUM(C4:C{tr-1})'
ws.cell(row=tr, column=4).value = f'=SUM(D4:D{tr-1})'
style_rows(ws, 4, tr, 5, money_cols=[3])
style_total(ws, tr, 5, money_cols=[3])
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 22

# --- PESTANA: Gastos Generales (TODOS los del CSV) ---
ws_gg = wb.create_sheet("Gastos Generales")
ws_gg.sheet_properties.tabColor = "E94560"
ws_gg['A1'] = f"GASTOS GENERALES ({len(gg_entries)} facturas)"
ws_gg['A1'].font = title_font
ws_gg.merge_cells('A1:F1')
for j, h in enumerate(["Codigo", "Fecha", "Titulo/Concepto", "Proveedor", "Importe (EUR)", "Observaciones"], 1):
    ws_gg.cell(row=3, column=j, value=h)
style_header(ws_gg, 3, 6)

for i, e in enumerate(gg_entries, start=4):
    ws_gg.cell(row=i, column=1, value=e['cod'])
    ws_gg.cell(row=i, column=2, value=e['fecha'])
    ws_gg.cell(row=i, column=3, value=e['titulo'])
    ws_gg.cell(row=i, column=4, value=e['proveedor'])
    ws_gg.cell(row=i, column=5, value=e['importe'])

tr_gg = 4 + len(gg_entries)
ws_gg.cell(row=tr_gg, column=3, value="TOTAL GASTOS GENERALES")
ws_gg.cell(row=tr_gg, column=5).value = f'=SUM(E4:E{tr_gg-1})'
style_rows(ws_gg, 4, tr_gg, 6, money_cols=[5])
style_total(ws_gg, tr_gg, 6, money_cols=[5])
ws_gg.column_dimensions['A'].width = 22
ws_gg.column_dimensions['B'].width = 14
ws_gg.column_dimensions['C'].width = 42
ws_gg.column_dimensions['D'].width = 30
ws_gg.column_dimensions['E'].width = 18
ws_gg.column_dimensions['F'].width = 18

# --- PESTANA: Vehiculos (TODOS los del CSV) ---
ws_vh = wb.create_sheet("Vehiculos")
ws_vh.sheet_properties.tabColor = "F39C12"
ws_vh['A1'] = f"GASTOS DE VEHICULOS ({len(vh_entries)} facturas)"
ws_vh['A1'].font = title_font
ws_vh.merge_cells('A1:F1')
for j, h in enumerate(["Codigo", "Fecha", "Titulo/Concepto", "Proveedor", "Importe (EUR)", "Observaciones"], 1):
    ws_vh.cell(row=3, column=j, value=h)
style_header(ws_vh, 3, 6)
for i, e in enumerate(vh_entries, start=4):
    ws_vh.cell(row=i, column=1, value=e['cod'])
    ws_vh.cell(row=i, column=2, value=e['fecha'])
    ws_vh.cell(row=i, column=3, value=e['titulo'])
    ws_vh.cell(row=i, column=4, value=e['proveedor'])
    ws_vh.cell(row=i, column=5, value=e['importe'])
tr_vh = 4 + len(vh_entries)
ws_vh.cell(row=tr_vh, column=3, value="TOTAL VEHICULOS")
ws_vh.cell(row=tr_vh, column=5).value = f'=SUM(E4:E{tr_vh-1})'
style_rows(ws_vh, 4, tr_vh, 6, money_cols=[5])
style_total(ws_vh, tr_vh, 6, money_cols=[5])
ws_vh.column_dimensions['A'].width = 22
ws_vh.column_dimensions['B'].width = 14
ws_vh.column_dimensions['C'].width = 42
ws_vh.column_dimensions['D'].width = 30
ws_vh.column_dimensions['E'].width = 18
ws_vh.column_dimensions['F'].width = 18

# --- PESTANA: Mano de Obra ---
ws_mo = wb.create_sheet("Mano de Obra")
ws_mo.sheet_properties.tabColor = "2ECC71"
ws_mo['A1'] = "GASTOS DE MANO DE OBRA"
ws_mo['A1'].font = title_font
ws_mo.merge_cells('A1:F1')
ws_mo['A2'] = "Tarifa: 20 EUR/hora (configurable)"
ws_mo['A2'].font = Font(bold=True, color='0F3460')
for j, h in enumerate(["Proyecto", "Horas", "Tarifa (EUR/h)", "Total (EUR)", "Ano", "% del Total"], 1):
    ws_mo.cell(row=4, column=j, value=h)
style_header(ws_mo, 4, 6)
for i, item in enumerate(mo_data, start=5):
    nombre, horas, anio = item[0], item[1], item[2]
    coste_directo = item[3] if len(item) > 3 else 0
    ws_mo.cell(row=i, column=1, value=map_mo_name(nombre))
    ws_mo.cell(row=i, column=2, value=horas)
    if horas > 0:
        ws_mo.cell(row=i, column=3, value=20)
        ws_mo.cell(row=i, column=4).value = f'=B{i}*C{i}'
    else:
        ws_mo.cell(row=i, column=3, value=0)
        ws_mo.cell(row=i, column=4, value=coste_directo)
    ws_mo.cell(row=i, column=5, value=anio)
tr_mo = 5 + len(mo_data)
ws_mo.cell(row=tr_mo, column=1, value="TOTAL MANO DE OBRA")
ws_mo.cell(row=tr_mo, column=2).value = f'=SUM(B5:B{tr_mo-1})'
ws_mo.cell(row=tr_mo, column=4).value = f'=SUM(D5:D{tr_mo-1})'
for i in range(5, tr_mo):
    ws_mo.cell(row=i, column=6).value = f'=D{i}/D{tr_mo}'
style_rows(ws_mo, 5, tr_mo, 6, money_cols=[3, 4])
style_total(ws_mo, tr_mo, 6, money_cols=[3, 4])
ws_mo.column_dimensions['A'].width = 35
ws_mo.column_dimensions['B'].width = 12
ws_mo.column_dimensions['C'].width = 14
ws_mo.column_dimensions['D'].width = 16
ws_mo.column_dimensions['E'].width = 10
ws_mo.column_dimensions['F'].width = 14

# --- PESTANA: Gastos Directos por Proyecto ---
ws_gd = wb.create_sheet("Gastos por Proyecto")
ws_gd.sheet_properties.tabColor = "9B59B6"
ws_gd['A1'] = "GASTOS DIRECTOS POR PROYECTO (del CSV de Holded)"
ws_gd['A1'].font = title_font
ws_gd.merge_cells('A1:C1')
for j, h in enumerate(["Proyecto", "Importe (EUR)", "Observaciones"], 1):
    ws_gd.cell(row=3, column=j, value=h)
style_header(ws_gd, 3, 3)
sorted_gd = sorted(gastos_directos.items(), key=lambda x: -x[1])
for i, (nombre, imp) in enumerate(sorted_gd, start=4):
    ws_gd.cell(row=i, column=1, value=nombre)  # Already mapped to cert name
    ws_gd.cell(row=i, column=2, value=imp)
tr_gd = 4 + len(sorted_gd)
ws_gd.cell(row=tr_gd, column=1, value="TOTAL GASTOS DIRECTOS")
ws_gd.cell(row=tr_gd, column=2).value = f'=SUM(B4:B{tr_gd-1})'
style_rows(ws_gd, 4, tr_gd, 3, money_cols=[2])
style_total(ws_gd, tr_gd, 3, money_cols=[2])
ws_gd.column_dimensions['A'].width = 50
ws_gd.column_dimensions['B'].width = 20
ws_gd.column_dimensions['C'].width = 22

# --- PESTANA: Facturas por Obra (todas las facturas directas del CSV) ---
ws_fo = wb.create_sheet("Facturas por Obra")
ws_fo.sheet_properties.tabColor = "E67E22"
ws_fo['A1'] = f"FACTURAS POR OBRA - Todas las facturas directas ({len(facturas_por_obra)} facturas)"
ws_fo['A1'].font = title_font
ws_fo.merge_cells('A1:G1')
ws_fo['A2'] = 'Usa los filtros desplegables en la fila 3 para buscar por proyecto, proveedor, etc.'
ws_fo['A2'].font = Font(italic=True, color='666666', size=10)
headers_fo = ["Proyecto", "Codigo", "Fecha", "Titulo/Concepto", "Proveedor", "Estado", "Importe (EUR)"]
for j, h in enumerate(headers_fo, 1):
    ws_fo.cell(row=3, column=j, value=h)
style_header(ws_fo, 3, 7)
# Sort by project then by amount (descending)
sorted_fo = sorted(facturas_por_obra, key=lambda x: (x['proyecto'], -abs(x['importe'])))
for i, e in enumerate(sorted_fo, start=4):
    ws_fo.cell(row=i, column=1, value=map_csv_to_cert(e['proyecto']))
    ws_fo.cell(row=i, column=2, value=e['cod'])
    ws_fo.cell(row=i, column=3, value=e['fecha'])
    ws_fo.cell(row=i, column=4, value=e['titulo'])
    ws_fo.cell(row=i, column=5, value=e['proveedor'])
    ws_fo.cell(row=i, column=6, value=e['estado'])
    ws_fo.cell(row=i, column=7, value=e['importe'])
    ws_fo.cell(row=i, column=7).number_format = money_format
    ws_fo.cell(row=i, column=7).border = thin_border
    for c in range(1, 7):
        ws_fo.cell(row=i, column=c).border = thin_border
tr_fo = 4 + len(sorted_fo)
ws_fo.cell(row=tr_fo, column=1, value="TOTAL FACTURAS POR OBRA")
ws_fo.cell(row=tr_fo, column=7).value = f'=SUM(G4:G{tr_fo-1})'
style_total(ws_fo, tr_fo, 7, money_cols=[7])
ws_fo.column_dimensions['A'].width = 50
ws_fo.column_dimensions['B'].width = 22
ws_fo.column_dimensions['C'].width = 14
ws_fo.column_dimensions['D'].width = 35
ws_fo.column_dimensions['E'].width = 30
ws_fo.column_dimensions['F'].width = 14
ws_fo.column_dimensions['G'].width = 18
# Auto-filter
ws_fo.auto_filter.ref = f'A3:G{tr_fo-1}'

# --- PESTANA: Resumen ---
ws_res = wb.create_sheet("Resumen")
ws_res.sheet_properties.tabColor = "16213E"
ws_res['A1'] = "RESUMEN FINANCIERO ECO STRUCT"
ws_res['A1'].font = title_font
ws_res.merge_cells('A1:I1')

ws_res['A3'] = "DATOS GENERALES"
ws_res['A3'].font = Font(bold=True, color='0F3460')
ws_res['A4'] = "Total Certificaciones:"
ws_res['B4'] = f"='Certificaciones'!C{tr}"
ws_res['A5'] = "Total Gastos Generales:"
ws_res['B5'] = f"='Gastos Generales'!E{tr_gg}"
ws_res['A6'] = "Total Vehiculos:"
ws_res['B6'] = f"='Vehiculos'!E{tr_vh}"
ws_res['A7'] = "Total Gastos Comunes:"
ws_res['B7'] = '=B5+B6'
ws_res['A8'] = "Total Mano de Obra:"
ws_res['B8'] = f"='Mano de Obra'!D{tr_mo}"
for r in range(4, 9):
    ws_res.cell(row=r, column=1).font = Font(bold=True)
    ws_res.cell(row=r, column=2).number_format = money_format
    ws_res.cell(row=r, column=2).border = thin_border

ws_res['A10'] = "DESGLLOSE POR PROYECTO (con Prorrateo)"
ws_res['A10'].font = Font(bold=True, color='0F3460')
headers = ["Proyecto", "Certificacion (EUR)", "% Certif.", "Gastos Prorrateados (EUR)", "Gastos Directos (EUR)", "Mano de Obra (EUR)", "TOTAL COSTE (EUR)", "Margen (EUR)", "% Coste/Certif."]
for j, h in enumerate(headers, 1):
    ws_res.cell(row=11, column=j, value=h)
style_header(ws_res, 11, 9)

cert_rows_range = range(4, 4 + len(cert_data))
res_rows_range = range(12, 12 + len(cert_data))
for i, (cert_r, res_r) in enumerate(zip(cert_rows_range, res_rows_range)):
    ws_res.cell(row=res_r, column=1).value = f"='Certificaciones'!B{cert_r}"
    ws_res.cell(row=res_r, column=2).value = f"='Certificaciones'!C{cert_r}"
    ws_res.cell(row=res_r, column=3).value = f"='Certificaciones'!D{cert_r}"
    ws_res.cell(row=res_r, column=4).value = f"=C{res_r}*$B$7"
    ws_res.cell(row=res_r, column=5).value = f"=SUMIF('Gastos por Proyecto'!A:A,A{res_r},'Gastos por Proyecto'!B:B)"
    ws_res.cell(row=res_r, column=6).value = f"=SUMIF('Mano de Obra'!A:A,A{res_r},'Mano de Obra'!D:D)"
    ws_res.cell(row=res_r, column=7).value = f"=D{res_r}+E{res_r}+F{res_r}"
    ws_res.cell(row=res_r, column=8).value = f"=B{res_r}-G{res_r}"
    ws_res.cell(row=res_r, column=9).value = f"=IF(B{res_r}=0,0,G{res_r}/B{res_r})"

tr_res = 12 + len(cert_data)
ws_res.cell(row=tr_res, column=1, value="TOTAL GENERAL")
for c in range(2, 9):
    cl = get_column_letter(c)
    ws_res.cell(row=tr_res, column=c).value = f"=SUM({cl}12:{cl}{tr_res-1})"
ws_res.cell(row=tr_res, column=9).value = f"=IF(B{tr_res}=0,0,G{tr_res}/B{tr_res})"
style_rows(ws_res, 12, tr_res, 9, money_cols=[2, 4, 5, 6, 7, 8])
style_total(ws_res, tr_res, 9, money_cols=[2, 4, 5, 6, 7, 8])
ws_res.column_dimensions['A'].width = 42
ws_res.column_dimensions['B'].width = 20
ws_res.column_dimensions['C'].width = 12
ws_res.column_dimensions['D'].width = 22
ws_res.column_dimensions['E'].width = 22
ws_res.column_dimensions['F'].width = 18
ws_res.column_dimensions['G'].width = 20
ws_res.column_dimensions['H'].width = 18
ws_res.column_dimensions['I'].width = 16

# --- PESTANA: Prorrateo ---
ws_pr = wb.create_sheet("Prorrateo")
ws_pr.sheet_properties.tabColor = "0F3460"
ws_pr['A1'] = "PRORRATEO DE GASTOS COMUNES"
ws_pr['A1'].font = title_font
ws_pr.merge_cells('A1:E1')
ws_pr['A3'] = "Formula: (% Certificacion del proyecto / Total) x Total Gastos Comunes"
ws_pr['A3'].font = Font(bold=True, color='0F3460')
for j, h in enumerate(["Proyecto", "Certificacion (EUR)", "% del Total", "Gasto Prorrateado (EUR)"], 1):
    ws_pr.cell(row=5, column=j, value=h)
style_header(ws_pr, 5, 4)
for i, cert_r in enumerate(cert_rows_range):
    r = 6 + i
    ws_pr.cell(row=r, column=1).value = f"='Certificaciones'!B{cert_r}"
    ws_pr.cell(row=r, column=2).value = f"='Certificaciones'!C{cert_r}"
    ws_pr.cell(row=r, column=3).value = f"='Certificaciones'!D{cert_r}"
    ws_pr.cell(row=r, column=4).value = f"=C{r}*'Resumen'!B7"
tr_pr = 6 + len(cert_data)
ws_pr.cell(row=tr_pr, column=1, value="TOTAL")
ws_pr.cell(row=tr_pr, column=2).value = f"=SUM(B6:B{tr_pr-1})"
ws_pr.cell(row=tr_pr, column=3).value = f"=SUM(C6:C{tr_pr-1})"
ws_pr.cell(row=tr_pr, column=4).value = f"=SUM(D6:D{tr_pr-1})"
style_rows(ws_pr, 6, tr_pr, 4, money_cols=[2, 4])
style_total(ws_pr, tr_pr, 4, money_cols=[2, 4])
ws_pr.column_dimensions['A'].width = 42
ws_pr.column_dimensions['B'].width = 20
ws_pr.column_dimensions['C'].width = 14
ws_pr.column_dimensions['D'].width = 22

# Mover Resumen al principio
wb.move_sheet("Resumen", offset=-6)

# ============================================================
# GUARDAR
# ============================================================
output = "ECO_STRUCT_Datos.xlsx"
wb.save(output)
print(f'\nExcel creado: {output}')
print(f'  - Certificaciones: {len(cert_data)} proyectos')
print(f'  - Gastos Generales: {len(gg_entries)} facturas = {gg_total:,.2f} EUR')
print(f'  - Vehiculos: {len(vh_entries)} facturas = {vh_total:,.2f} EUR')
print(f'  - Total Gastos Comunes: {gg_total + vh_total:,.2f} EUR')
print(f'  - Mano de Obra: {len(mo_data)} registros')
print(f'  - Gastos Directos: {len(gastos_directos)} proyectos')
print(f'  - Facturas por Obra: {len(facturas_por_obra)} facturas')
